# main.py
import os
import asyncio
from io import BytesIO
from datetime import datetime
from dotenv import load_dotenv
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, BotCommand, InputFile
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, CallbackQueryHandler, filters, ContextTypes

import config
import database
import json
import re

# load .env
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DOTENV_PATH = os.path.join(BASE_DIR, '.env')
if os.path.exists(DOTENV_PATH):
    load_dotenv(dotenv_path=DOTENV_PATH)
else:
    load_dotenv()

TOKEN = os.getenv('BOT_TOKEN')
if not TOKEN:
    print("ERROR: BOT_TOKEN not found in env (BOT_TOKEN)")

user_states = {}


def sync_runtime_config():
    """Load dynamic lists (questions/RTP/FCKP options) from DB settings if present."""
    # Questions for MKK report
    try:
        qs = database.get_mkk_questions()
        if qs:
            config.QUESTIONS = [{"key": q["key"], "question": q["question"]} for q in qs]
    except Exception:
        pass

    # RTP list
    try:
        rtps = database.get_rtp_list()
        if rtps:
            config.RTP_LIST = rtps
    except Exception:
        pass

    # FCKP/CKP options (buttons)
    try:
        raw = database.get_setting("fckp_options")
        if raw:
            opts = json.loads(raw)
            if isinstance(opts, list):
                opts = [str(x).strip() for x in opts if str(x).strip()]
                if opts:
                    config.FCKP_OPTIONS = opts
    except Exception:
        pass


def safe_state(uid):
    st = user_states.get(uid)
    if not st:
        st = {'mode': 'idle', 'step': 0, 'data': {}, 'editing': False}
        user_states[uid] = st
    return st


def build_main_menu():
    kb = [
        [InlineKeyboardButton("👥 Отчет МКК", callback_data='role_mkk')],
        [InlineKeyboardButton("👤 РТП", callback_data='role_rtp')],
        [InlineKeyboardButton("🏢 УПМБ", callback_data='role_rm')],
        [InlineKeyboardButton("🛠 Администрирование", callback_data='role_admin')],
        [InlineKeyboardButton("Сменить ФИ/РТП", callback_data='change_info')]
    ]
    return InlineKeyboardMarkup(kb)


def sanitize_sheet_title(title: str) -> str:
    """Excel sheet titles must be <=31 chars and cannot contain : \ / ? * [ ]"""
    title = re.sub(r'[:\\/?*\[\]]', '_', str(title or '').strip())
    title = title.strip() or "Sheet1"
    return title[:31]


def sanitize_filename(name: str, default_base: str = "report") -> str:
    """Make a safe filename for Telegram documents."""
    base = re.sub(r'[^0-9A-Za-zА-Яа-яЁё._-]+', '_', str(name or '')).strip('._-')
    if not base:
        base = default_base
    return base[:120]


async def send_or_edit(target, text: str, reply_markup=None):
    """Edit a message if possible (CallbackQuery), otherwise send a new one."""
    # CallbackQuery-like
    try:
        if hasattr(target, "edit_message_text"):
            return await target.edit_message_text(text, reply_markup=reply_markup)
    except Exception:
        pass
    # Message-like
    try:
        if hasattr(target, "reply_text"):
            return await target.reply_text(text, reply_markup=reply_markup)
    except Exception:
        pass
    # CallbackQuery.message fallback
    try:
        if hasattr(target, "message") and target.message:
            return await target.message.reply_text(text, reply_markup=reply_markup)
    except Exception:
        return None


# --- GOALS helpers ---

def _today_iso() -> str:
    return datetime.now().strftime('%Y-%m-%d')


def _iso_to_ru(d: str) -> str:
    try:
        return datetime.strptime(d, '%Y-%m-%d').strftime('%d.%m.%Y')
    except Exception:
        return str(d or '')


def _parse_date_to_iso(s: str) -> str:
    s = (s or '').strip()
    if not s:
        raise ValueError('empty')
    low = s.lower()
    if low in ('сегодня', 'today'):
        return _today_iso()
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except Exception:
            pass
    raise ValueError('bad date format')


def _to_float(v) -> float:
    try:
        if v is None or v == '':
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        return float(str(v).replace(',', '.').strip())
    except Exception:
        return 0.0


def _metric_label(metric_type: str, metric_key: str) -> str:
    metric_type = (metric_type or '').lower()
    if metric_type == 'question':
        for q in getattr(config, 'QUESTIONS', []):
            if q.get('key') == metric_key:
                return (q.get('question') or metric_key).strip()
        return str(metric_key)
    if metric_type == 'fckp_total':
        return 'ФЦКП (всего)'
    if metric_type == 'fckp_product':
        return f"ФЦКП: {metric_key}"
    return str(metric_key)


def _compute_goal_achieved(goal: dict, today_iso: str = None) -> float:
    today_iso = today_iso or _today_iso()
    date_from = goal.get('date_from') or today_iso
    date_to = goal.get('date_to') or today_iso
    end = min(today_iso, date_to)
    if end < date_from:
        return 0.0

    scope = (goal.get('scope') or '').lower()
    owner = goal.get('owner_name')
    metric_type = (goal.get('metric_type') or '').lower()
    metric_key = goal.get('metric_key')

    try:
        rows = database.get_mkk_reports_between(date_from, end)
    except Exception:
        rows = []

    achieved = 0.0
    for uid, rdate, data, current_mfi in rows:
        if not isinstance(data, dict):
            continue
        if scope == 'team':
            snap = data.get('manager_fi_snapshot') or current_mfi
            if snap != owner:
                continue

        if metric_type == 'question':
            achieved += _to_float(data.get(metric_key, 0))
        elif metric_type == 'fckp_total':
            prods = data.get('fckp_products')
            if isinstance(prods, list):
                achieved += float(len(prods))
            else:
                achieved += _to_float(data.get('fckp_realized', 0))
        elif metric_type == 'fckp_product':
            prods = data.get('fckp_products')
            if isinstance(prods, list):
                achieved += float(sum(1 for p in prods if str(p) == str(metric_key)))
    return achieved


def _compute_goal_user_scores(goal: dict, today_iso: str = None) -> dict:
    """Return dict {user_id: achieved} for the goal period up to today."""
    today_iso = today_iso or _today_iso()
    date_from = goal.get('date_from') or today_iso
    date_to = goal.get('date_to') or today_iso
    end = min(today_iso, date_to)
    if end < date_from:
        return {}

    scope = (goal.get('scope') or '').lower()
    owner = goal.get('owner_name')
    metric_type = (goal.get('metric_type') or '').lower()
    metric_key = goal.get('metric_key')

    try:
        rows = database.get_mkk_reports_between(date_from, end)
    except Exception:
        rows = []

    scores = {}
    for uid, rdate, data, current_mfi in rows:
        if not isinstance(data, dict):
            continue
        if scope == 'team':
            snap = data.get('manager_fi_snapshot') or current_mfi
            if snap != owner:
                continue

        add = 0.0
        if metric_type == 'question':
            add = _to_float(data.get(metric_key, 0))
        elif metric_type == 'fckp_total':
            prods = data.get('fckp_products')
            if isinstance(prods, list):
                add = float(len(prods))
            else:
                add = _to_float(data.get('fckp_realized', 0))
        elif metric_type == 'fckp_product':
            prods = data.get('fckp_products')
            if isinstance(prods, list):
                add = float(sum(1 for p in prods if str(p) == str(metric_key)))

        if add > 0:
            scores[int(uid)] = scores.get(int(uid), 0.0) + float(add)

    return scores


def _format_goal_leaderboard_lines(goal: dict, top_n: int, today_iso: str = None) -> list:
    """Return formatted leaderboard lines for a goal (only users with >0 progress)."""
    top_n = int(top_n or 0)
    if top_n <= 0:
        return []

    scores = _compute_goal_user_scores(goal, today_iso=today_iso)
    items = [(uid, val) for uid, val in scores.items() if float(val) > 0]
    if not items:
        return []

    items.sort(key=lambda x: (-float(x[1]), int(x[0])))
    items = items[:top_n]

    try:
        names = database.get_user_names_by_ids([uid for uid, _ in items])
    except Exception:
        names = {}

    medals = ['🥇', '🥈', '🥉']
    lines = []
    for i, (uid, val) in enumerate(items):
        icon = medals[i] if i < 3 else '🔥'
        fio = names.get(int(uid)) or str(uid)
        lines.append(f"  {icon} {fio} — {config.format_value(val)}")
    return lines


def _format_goal_short(goal: dict, achieved: float) -> str:
    target = _to_float(goal.get('target_value', 0))
    remaining = max(0.0, target - achieved)
    due = _iso_to_ru(goal.get('date_to'))
    title = (goal.get('title') or '').strip()
    a = config.format_value(achieved)
    t = config.format_value(target)
    r = config.format_value(remaining)
    return f"• {title}: {a}/{t} (осталось {r}) до {due}"


def _start_goals_block(uid: int) -> str:
    today = _today_iso()
    try:
        database.cleanup_expired_goals(today)
    except Exception:
        pass

    lines = []

    try:
        gosb = database.list_goals('gosb', today=today)
    except Exception:
        gosb = []

    if gosb:
        lines.append('🎯 Цели ГОСБ:')
        for g in gosb[:3]:
            achieved = _compute_goal_achieved(g, today)
            lines.append(_format_goal_short(g, achieved))
            # Optional TOP employees for this goal
            try:
                top_n = database.get_goal_leaderboard_top_n(int(g.get('id')))
            except Exception:
                top_n = 0
            lb_lines = _format_goal_leaderboard_lines(g, top_n, today_iso=today) if top_n else []
            if lb_lines:
                lines.extend(lb_lines)
        if len(gosb) > 3:
            lines.append(f"… и ещё {len(gosb) - 3} целей")

    owner = None
    try:
        role = database.get_user_role(uid)
    except Exception:
        role = None

    try:
        if role == 'rtp':
            owner = database.get_user_name(uid)
        else:
            owner = database.get_manager_fi_for_employee(uid)
    except Exception:
        owner = None

    if owner:
        try:
            team = database.list_goals('team', owner_name=owner, today=today)
        except Exception:
            team = []
        if team:
            lines.append('')
            lines.append(f"👥 Цели команды ({owner}):")
            for g in team[:3]:
                achieved = _compute_goal_achieved(g, today)
                lines.append(_format_goal_short(g, achieved))
                # Optional TOP employees for this goal
                try:
                    top_n = database.get_goal_leaderboard_top_n(int(g.get('id')))
                except Exception:
                    top_n = 0
                lb_lines = _format_goal_leaderboard_lines(g, top_n, today_iso=today) if top_n else []
                if lb_lines:
                    lines.extend(lb_lines)
            if len(team) > 3:
                lines.append(f"… и ещё {len(team) - 3} целей")

    return '\n'.join(lines).strip()


def _metric_picker_keyboard():
    kb = []
    for q in getattr(config, 'QUESTIONS', []):
        label = (q.get('question') or q.get('key') or '').strip()[:64]
        kb.append([InlineKeyboardButton(label, callback_data=f"goal_metric_q_{q.get('key')}")])
    kb.append([InlineKeyboardButton('ФЦКП (всего)', callback_data='goal_metric_fckp_total')])
    for p in getattr(config, 'FCKP_OPTIONS', []):
        kb.append([InlineKeyboardButton(f"ФЦКП: {p}", callback_data=f"goal_metric_fckp_prod_{p}")])
    kb.append([InlineKeyboardButton('⬅️ Назад', callback_data='goal_cancel_metric')])
    return InlineKeyboardMarkup(kb)


# --- Helpers for xlsx generation (used by RM) ---
def generate_xlsx_for_report(title: str, rows: list, columns: list):
    """Generate an .xlsx file in memory (BytesIO)."""
    try:
        import openpyxl
    except Exception as e:
        raise RuntimeError("openpyxl не установлен. Установите: pip install openpyxl") from e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sanitize_sheet_title(title)

    # Header
    for c_idx, (_, col_title) in enumerate(columns, start=1):
        ws.cell(row=1, column=c_idx, value=col_title)

    # Rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (col_key, _) in enumerate(columns, start=1):
            value = row.get(col_key, "") if isinstance(row, dict) else ""
            # Keep lists/dicts readable
            if isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False)
            ws.cell(row=r_idx, column=c_idx, value=value)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# --- Handlers ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    sync_runtime_config()
    msg = update.message or update.effective_message
    block = _start_goals_block(msg.from_user.id)
    text = (block + "\n\n" if block else "") + "Выберите роль:"
    await msg.reply_text(text, reply_markup=build_main_menu())


async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not query:
        return
    await query.answer()
    uid = query.from_user.id
    data = query.data or ""
    st = user_states.get(uid, {})
    sync_runtime_config()

    # GOAL callbacks
    if data == 'goal_cancel_metric':
        gs = (st or {}).get('goal_scope')
        if gs == 'team':
            await show_goals_menu(query, uid, scope='team', owner_name=(st or {}).get('goal_owner'), back_cb='rtp_menu')
        else:
            await show_goals_menu(query, uid, scope='gosb', back_cb='rm_management')
        return

    if data.startswith('goal_metric_'):
        st2 = safe_state(uid)
        if st2.get('mode') != 'goal_pick_metric':
            return
        metric_type = None
        metric_key = None
        if data.startswith('goal_metric_q_'):
            metric_type = 'question'
            metric_key = data.split('goal_metric_q_', 1)[1]
        elif data == 'goal_metric_fckp_total':
            metric_type = 'fckp_total'
            metric_key = 'fckp_total'
        elif data.startswith('goal_metric_fckp_prod_'):
            metric_type = 'fckp_product'
            metric_key = data.split('goal_metric_fckp_prod_', 1)[1]

        if not metric_type or not metric_key:
            await send_or_edit(query, 'Ошибка выбора показателя.')
            return

        action = st2.get('goal_action')
        scope = st2.get('goal_scope')
        owner = st2.get('goal_owner')

        if action == 'add':
            st2['goal_metric_type'] = metric_type
            st2['goal_metric_key'] = metric_key
            st2['mode'] = 'goal_add_target'
            await send_or_edit(query, 'Введите целевое значение (число):', reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton('⬅️ Назад', callback_data=f"{scope}_goals_menu")]]))
            return

        if action == 'edit_metric':
            goal_id = st2.get('goal_id')
            try:
                database.update_goal(int(goal_id), metric_type=metric_type, metric_key=metric_key)
            except Exception:
                pass
            await show_goal_edit_menu(query, uid, scope=scope, goal_id=goal_id, owner_name=owner)
            return
        return

    if data.startswith('goal_add_'):
        scope = data.split('_', 2)[2]
        st2 = safe_state(uid)
        st2.clear()
        st2.update({'mode': 'goal_add_title', 'goal_scope': scope, 'editing': False})
        if scope == 'team':
            st2['goal_owner'] = database.get_user_name(uid)
        await send_or_edit(query, 'Введите название цели:', reply_markup=InlineKeyboardMarkup(
            [[InlineKeyboardButton('⬅️ Назад', callback_data=f"{scope}_goals_menu")]]))
        return

    if data.startswith('goal_edit_') and data.count('_') == 3:
        _, _, scope, gid = data.split('_', 3)
        st2 = safe_state(uid)
        st2['goal_scope'] = scope
        st2['goal_owner'] = database.get_user_name(uid) if scope == 'team' else None
        await show_goal_edit_menu(query, uid, scope=scope, goal_id=int(gid), owner_name=st2.get('goal_owner'))
        return

    if data.startswith('goal_editfield_'):
        parts = data.split('_')
        if len(parts) < 5:
            return
        scope = parts[2]
        gid = int(parts[3])
        field = parts[4]
        st2 = safe_state(uid)
        st2['goal_scope'] = scope
        st2['goal_id'] = gid
        if scope == 'team':
            st2['goal_owner'] = database.get_user_name(uid)

        if field == 'metric':
            st2['mode'] = 'goal_pick_metric'
            st2['goal_action'] = 'edit_metric'
            await send_or_edit(query, 'Выберите показатель:', reply_markup=_metric_picker_keyboard())
            return

        if field in ('title', 'target', 'date_from', 'date_to'):
            st2['mode'] = f"goal_edit_{field}"
            prompt = {
                'title': 'Введите новое название цели:',
                'target': 'Введите новое целевое значение (число):',
                'date_from': 'Введите новую дату начала (ДД.ММ.ГГГГ) или YYYY-MM-DD:',
                'date_to': 'Введите новую дату окончания (ДД.ММ.ГГГГ) или YYYY-MM-DD:',
            }.get(field, 'Введите значение:')
            await send_or_edit(query, prompt, reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton('⬅️ Назад', callback_data=f"goal_edit_{scope}_{gid}")]]))
            return
        return

    if data.startswith('goal_del_'):
        parts = data.split('_')
        if len(parts) < 4:
            return
        scope = parts[2]
        gid = int(parts[3])
        g = None
        try:
            g = database.get_goal(gid)
        except Exception:
            g = None
        title = (g or {}).get('title') or f"#{gid}"
        kb = [[
            InlineKeyboardButton('✅ Да, удалить', callback_data=f"goal_delconfirm_{scope}_{gid}"),
            InlineKeyboardButton('⬅️ Нет', callback_data=f"goal_edit_{scope}_{gid}")
        ]]
        await send_or_edit(query, f"Удалить цель {title}?", reply_markup=InlineKeyboardMarkup(kb))
        return

    if data.startswith('goal_delconfirm_'):
        parts = data.split('_')
        if len(parts) < 4:
            return
        scope = parts[2]
        gid = int(parts[3])
        try:
            database.delete_goal(gid)
        except Exception:
            pass
        if scope == 'team':
            owner = database.get_user_name(uid)
            await show_goals_menu(query, uid, scope='team', owner_name=owner, back_cb='rtp_menu')
        else:
            await show_goals_menu(query, uid, scope='gosb', back_cb='rm_management')
        return

    # return to main
    if data == 'return_to_menu':
        user_states.pop(uid, None)
        await query.edit_message_text("Выберите роль:", reply_markup=build_main_menu())
        return

    # -----------------------------
    # ADMIN callbacks
    # -----------------------------
    if data == 'admin_menu':
        user_states[uid] = {'mode': 'admin', 'step': 0, 'data': {}, 'editing': False}
        await show_admin_menu(query)
        return

    if data == 'admin_edit_questions':
        user_states[uid] = {'mode': 'admin_edit_questions'}
        await show_admin_questions_editor(query)
        return
    if data == 'admin_edit_fckp':
        user_states[uid] = {'mode': 'admin_edit_fckp'}
        await show_admin_fckp_editor(query)
        return

    if data == 'admin_fckp_add':
        user_states[uid] = {'mode': 'admin_fckp_add'}
        await send_or_edit(query, "Введите название новой кнопки ЦКП:", reply_markup=InlineKeyboardMarkup(
            [[InlineKeyboardButton("⬅️ Назад", callback_data='admin_edit_fckp')]]))
        return

    if data.startswith('admin_fckp_edit_'):
        try:
            idx = int(data.split('_')[-1])
        except Exception:
            await send_or_edit(query, "Ошибка выбора элемента.")
            return
        user_states[uid] = {'mode': 'admin_fckp_edit', 'fckp_idx': idx}
        await send_or_edit(query, "Введите новое название кнопки ЦКП:", reply_markup=InlineKeyboardMarkup(
            [[InlineKeyboardButton("⬅️ Назад", callback_data='admin_edit_fckp')]]))
        return

    if data.startswith('admin_fckp_del_'):
        try:
            idx = int(data.split('_')[-1])
        except Exception:
            await send_or_edit(query, "Ошибка удаления.")
            return
        opts = get_fckp_options()
        if 0 <= idx < len(opts):
            opts.pop(idx)
            save_fckp_options(opts)
        await show_admin_fckp_editor(query)
        return

    if data.startswith('admin_fckp_up_') or data.startswith('admin_fckp_down_'):
        try:
            parts = data.split('_')
            direction = parts[2]
            idx = int(parts[3])
        except Exception:
            await send_or_edit(query, "Ошибка перемещения.")
            return
        opts = get_fckp_options()
        if direction == 'up' and 0 < idx < len(opts):
            opts[idx - 1], opts[idx] = opts[idx], opts[idx - 1]
            save_fckp_options(opts)
        if direction == 'down' and 0 <= idx < len(opts) - 1:
            opts[idx + 1], opts[idx] = opts[idx], opts[idx + 1]
            save_fckp_options(opts)
        await show_admin_fckp_editor(query)
        return

    if data == 'admin_questions_add':
        user_states[uid] = {'mode': 'admin_questions_add'}
        try:
            await query.edit_message_text("Введите текст нового вопроса для отчёта МКК:")
        except Exception:
            await query.message.reply_text("Введите текст нового вопроса для отчёта МКК:")
        return

    if data.startswith('admin_q_edit_'):
        q_key = data.split('admin_q_edit_', 1)[1]
        user_states[uid] = {'mode': 'admin_q_edit', 'q_key': q_key}
        # покажем текущий текст
        cur = None
        try:
            for q in database.get_mkk_questions():
                if q.get('key') == q_key:
                    cur = q.get('question')
                    break
        except Exception:
            pass
        try:
            await query.edit_message_text(f"Текущий текст:\n{cur or ''}\n\nВведите новый текст вопроса:")
        except Exception:
            await query.message.reply_text(f"Текущий текст:\n{cur or ''}\n\nВведите новый текст вопроса:")
        return

    if data.startswith('admin_q_del_'):
        q_key = data.split('admin_q_del_', 1)[1]
        try:
            database.delete_mkk_question(q_key)
        except Exception:
            pass
        sync_runtime_config()
        await show_admin_questions_editor(query)
        return

    if data.startswith('admin_q_up_'):
        q_key = data.split('admin_q_up_', 1)[1]
        try:
            database.move_mkk_question(q_key, 'up')
        except Exception:
            pass
        sync_runtime_config()
        await show_admin_questions_editor(query)
        return

    if data.startswith('admin_q_down_'):
        q_key = data.split('admin_q_down_', 1)[1]
        try:
            database.move_mkk_question(q_key, 'down')
        except Exception:
            pass
        sync_runtime_config()
        await show_admin_questions_editor(query)
        return

    if data == 'admin_edit_rtps':
        user_states[uid] = {'mode': 'admin_edit_rtps'}
        await show_admin_rtps_editor(query)
        return

    # admin: employees editor
    if data == 'admin_emp_editor':
        user_states[uid] = {'mode': 'admin_emp_editor'}
        await show_admin_employees_rtp_selector(query)
        return

    if data.startswith('admin_emp_rtp_'):
        try:
            rtp_idx = int(data.split('_')[3])
        except Exception:
            await send_or_edit(query, "Ошибка выбора РТП.")
            return
        await show_admin_employees_list(query, rtp_idx)
        return

    if data.startswith('admin_emp_edit_'):
        # admin_emp_edit_{rtp_idx}_{emp_id}
        try:
            parts = data.split('_')
            rtp_idx = int(parts[3])
            emp_id = int(parts[4])
        except Exception:
            await send_or_edit(query, "Ошибка выбора сотрудника.")
            return
        await show_admin_employee_reassign(query, rtp_idx, emp_id)
        return

    if data.startswith('admin_emp_set_'):
        # admin_emp_set_{from_rtp_idx}_{emp_id}_{new_rtp_idx}
        try:
            parts = data.split('_')
            from_idx = int(parts[3])
            emp_id = int(parts[4])
            new_idx = int(parts[5])
        except Exception:
            await send_or_edit(query, "Ошибка назначения.")
            return
        rtps = database.get_rtp_list() or list(config.RTP_LIST)
        if new_idx < 0 or new_idx >= len(rtps):
            await send_or_edit(query, "Некорректный индекс РТП.")
            return
        database.set_manager_fi_for_employee(emp_id, rtps[new_idx])
        await show_admin_employees_list(query, from_idx)
        return

    if data.startswith('admin_emp_unbind_'):
        # admin_emp_unbind_{from_rtp_idx}_{emp_id}
        try:
            parts = data.split('_')
            from_idx = int(parts[3])
            emp_id = int(parts[4])
        except Exception:
            await send_or_edit(query, "Ошибка отвязки.")
            return
        database.set_manager_fi_for_employee(emp_id, None)
        await show_admin_employees_list(query, from_idx)
        return

    if data.startswith('admin_emp_del_'):
        # admin_emp_del_{from_rtp_idx}_{emp_id}
        try:
            parts = data.split('_')
            from_idx = int(parts[3])
            emp_id = int(parts[4])
        except Exception:
            await send_or_edit(query, "Ошибка удаления.")
            return
        try:
            database.delete_user(emp_id)
        except Exception as e:
            await send_or_edit(query, f"Ошибка удаления: {e}")
            return
        await show_admin_employees_list(query, from_idx)
        return

    if data == 'admin_rtp_add':
        user_states[uid] = {'mode': 'admin_rtp_add'}
        try:
            await query.edit_message_text("Введите ФИ нового РТП:")
        except Exception:
            await query.message.reply_text("Введите ФИ нового РТП:")
        return

    if data.startswith('admin_rtp_edit_'):
        idx_s = data.split('admin_rtp_edit_', 1)[1]
        try:
            idx = int(idx_s)
        except Exception:
            idx = -1
        rtps = []
        try:
            rtps = database.get_rtp_list()
        except Exception:
            rtps = getattr(config, 'RTP_LIST', [])
        if idx < 0 or idx >= len(rtps):
            await query.edit_message_text("Некорректный индекс РТП.")
            return
        old_name = rtps[idx]
        user_states[uid] = {'mode': 'admin_rtp_edit', 'old_name': old_name}
        user_states[uid] = {'mode': 'admin_rtp_edit', 'old_name': old_name}
        try:
            await query.edit_message_text(f"Текущее ФИ РТП:\n{old_name}\n\nВведите новое ФИ:")
        except Exception:
            await query.message.reply_text(f"Текущее ФИ РТП:\n{old_name}\n\nВведите новое ФИ:")
        return

    if data.startswith('admin_rtp_del_'):
        idx_s = data.split('admin_rtp_del_', 1)[1]
        try:
            idx = int(idx_s)
        except Exception:
            idx = -1
        rtps = []
        try:
            rtps = database.get_rtp_list()
        except Exception:
            rtps = getattr(config, 'RTP_LIST', [])
        if idx < 0 or idx >= len(rtps):
            await query.edit_message_text("Некорректный индекс РТП.")
            return
        try:
            database.delete_rtp(rtps[idx])
        except Exception:
            pass
        sync_runtime_config()
        await show_admin_rtps_editor(query)
        return

    if data.startswith('admin_rtp_up_'):
        idx_s = data.split('admin_rtp_up_', 1)[1]
        try:
            idx = int(idx_s)
        except Exception:
            idx = -1
        rtps = []
        try:
            rtps = database.get_rtp_list()
        except Exception:
            rtps = getattr(config, 'RTP_LIST', [])
        if idx < 0 or idx >= len(rtps):
            await query.edit_message_text("Некорректный индекс РТП.")
            return
        try:
            database.move_rtp(rtps[idx], 'up')
        except Exception:
            pass
        sync_runtime_config()
        await show_admin_rtps_editor(query)
        return

    if data.startswith('admin_rtp_down_'):
        idx_s = data.split('admin_rtp_down_', 1)[1]
        try:
            idx = int(idx_s)
        except Exception:
            idx = -1
        rtps = []
        try:
            rtps = database.get_rtp_list()
        except Exception:
            rtps = getattr(config, 'RTP_LIST', [])
        if idx < 0 or idx >= len(rtps):
            await query.edit_message_text("Некорректный индекс РТП.")
            return
        try:
            database.move_rtp(rtps[idx], 'down')
        except Exception:
            pass
        sync_runtime_config()
        await show_admin_rtps_editor(query)
        return

    if data == 'admin_set_rtp_password':
        user_states[uid] = {'mode': 'admin_set_rtp_password'}
        try:
            await query.edit_message_text("Введите новый пароль для входа РТП:")
        except Exception:
            await query.message.reply_text("Введите новый пароль для входа РТП:")
        return

    # role selection (common)
    if data.startswith('role_'):
        role = data.split('_', 1)[1]

        # Админ-панель (тот же пароль, что и для руководителей)
        if role == 'admin':
            if database.is_user_verified(uid):
                user_states[uid] = {'mode': 'admin', 'step': 0, 'data': {}, 'editing': False}
                await show_admin_menu(query)
                return
            user_states[uid] = {'mode': 'awaiting_admin_password'}
            try:
                await query.edit_message_text("Введите пароль администратора:")
            except Exception:
                await query.message.reply_text("Введите пароль администратора:")
            return

        # РМ/МН: пароль руководителя (ADMIN_PASSWORD)
        if role == 'rm':
            if database.is_user_verified(uid):
                user_states[uid] = {'mode': 'rm', 'step': 0, 'data': {}, 'editing': False}
                await handle_role_selection(query, uid, 'rm')
                return
            user_states[uid] = {'mode': 'awaiting_password_for', 'await_role': 'rm'}
            try:
                await query.edit_message_text("Введите пароль для доступа в раздел руководителя:")
            except Exception:
                await query.message.reply_text("Введите пароль для доступа в раздел руководителя:")
            return

        # РТП: отдельный пароль РТП (хранится в БД, меняется админом)
        if role == 'rtp':
            try:
                verified = database.is_user_rtp_verified(uid)
            except Exception:
                verified = database.is_user_verified(uid)

            if verified:
                user_states[uid] = {'mode': 'rtp', 'step': 0, 'data': {}, 'editing': False}
                await handle_role_selection(query, uid, 'rtp')
                return

            user_states[uid] = {'mode': 'awaiting_rtp_password'}
            try:
                await query.edit_message_text("Введите пароль для доступа в раздел РТП:")
            except Exception:
                await query.message.reply_text("Введите пароль для доступа в раздел РТП:")
            return

        # МКК
        if role == 'mkk':
            user_states[uid] = {'mode': 'mkk', 'step': 0, 'data': {}, 'editing': False}
            await handle_role_selection(query, uid, 'mkk')
            return

        # fallback
        user_states[uid] = {'mode': 'idle', 'step': 0, 'data': {}, 'editing': False}
        await query.edit_message_text("Неизвестная роль. Нажмите /start.")
        return

    # change_info
    if data == 'change_info':
        user_states[uid] = {'mode': 'change_fi_enter_name'}
        try:
            await query.edit_message_text("Введите ваше ФИ (как хотите, чтобы оно сохранялось):")
        except Exception:
            await query.message.reply_text("Введите ваше ФИ (как хотите, чтобы оно сохранялось):")
        return

    # choose_rtp_{idx}
    if data.startswith('choose_rtp_'):
        try:
            idx = int(data.split('_')[2])
        except Exception:
            await query.edit_message_text("Ошибка выбора. Попробуйте снова.")
            return
        if idx < 0 or idx >= len(config.RTP_LIST):
            await query.edit_message_text("Некорректный индекс РТП.")
            return
        selected = config.RTP_LIST[idx]
        # if in change_flow (user entered new name earlier)
        if st.get('change_flow'):
            new_name = st.get('new_name')
            if not new_name:
                await query.edit_message_text("Ошибка: имя не найдено в состоянии.")
                return
            database.add_user(uid, 'mkk', new_name, selected)
            user_states.pop(uid, None)
            await query.edit_message_text(f"Готово. Ваше имя '{new_name}' привязано к РТП: {selected}.")
            return

        role = st.get('mode', 'idle')
        if role == 'rtp':
            # user choosing their own FI as RTP
            database.add_user(uid, 'rtp', selected)
            # when RTP chooses own FI, ensure RTP password version remembered
            try:
                database.set_user_rtp_verified_version(uid, database.get_rtp_password_version())
            except Exception:
                database.set_user_verified(uid, 1)
            user_states[uid] = {'mode': 'rtp', 'step': 0, 'data': {}, 'editing': False}
            await query.edit_message_text(f"Вы вошли как РТП: {selected}")
            await show_manager_menu(query)
            return

        # registration flow for MKK
        name = st.get('name')
        if name:
            database.add_user(uid, 'mkk', name, selected)
            st.pop('choosing_rtp', None);
            st.pop('name', None)
            st.update({'step': 0, 'data': {}, 'editing': False, 'mode': 'mkk'})
            await query.edit_message_text(f"Привязка к {selected} успешна. Начинаем отчёт.")
            await ask_next_question(query.message, uid)
            return

        await query.edit_message_text("Непонятный контекст выбора РТП.")
        return

    # choose_rm_{idx} - RM selects their FI from list
    if data.startswith('choose_rm_'):
        try:
            idx = int(data.split('_')[2])
        except Exception:
            await query.edit_message_text("Ошибка выбора РМ/МН.")
            return
        if idx < 0 or idx >= len(config.RM_MN_LIST):
            await query.edit_message_text("Некорректный индекс.")
            return
        chosen = config.RM_MN_LIST[idx]
        # register user as rm and mark verified
        database.add_user(uid, 'rm', chosen)
        database.set_user_verified(uid, 1)
        user_states[uid] = {'mode': 'rm', 'step': 0, 'data': {}, 'editing': False}
        await show_rm_home(query, uid)
        return

    # role_rm -> show RM menu (entry)
    if data == 'role_rm':
        if database.is_user_verified(uid):
            await handle_role_selection(query, uid, 'rm')
            return
        else:
            user_states[uid] = {'mode': 'awaiting_password_for', 'await_role': 'rm'}
            try:
                await query.edit_message_text("Введите пароль для доступа в раздел руководителя:")
            except Exception:
                await query.message.reply_text("Введите пароль для доступа в раздел руководителя:")
            return

    if data == 'rm_menu':
        await show_rm_home(query, uid)
        return

    if data == 'rm_management':
        await show_rm_management_menu(query, uid)
        return

    if data == 'gosb_goals_menu':
        await show_goals_menu(query, uid, scope='gosb', back_cb='rm_management')
        return

    if data == 'gosb_leaderboards_menu':
        await show_leaderboards_menu(query, uid, scope='gosb', back_cb='rm_management')
        return

    if data == 'team_leaderboards_menu':
        try:
            owner = database.get_user_name(uid)
        except Exception:
            owner = None
        if not owner:
            await send_or_edit(query, 'Не удалось определить РТП для целей команды.')
            return
        await show_leaderboards_menu(query, uid, scope='team', owner_name=owner, back_cb='rtp_menu')
        return

    if data.startswith('lb_cfg_'):
        # lb_cfg_{scope}_{goal_id}
        try:
            parts = data.split('_')
            scope = parts[2]
            gid = int(parts[3])
        except Exception:
            await send_or_edit(query, 'Ошибка выбора цели.')
            return
        owner = None
        if scope == 'team':
            try:
                owner = database.get_user_name(uid)
            except Exception:
                owner = None
            if not owner:
                await send_or_edit(query, 'Не удалось определить РТП.')
                return
            # ensure goal belongs to this RTP
            try:
                g = database.get_goal(gid)
            except Exception:
                g = None
            if not g or (g.get('owner_name') != owner):
                await send_or_edit(query, 'Цель недоступна.')
                return
        await show_leaderboard_goal_config(query, uid, scope=scope, goal_id=gid, owner_name=owner)
        return

    if data.startswith('lb_setn_'):
        # lb_setn_{scope}_{goal_id}_{n}
        try:
            parts = data.split('_')
            scope = parts[2]
            gid = int(parts[3])
            n = int(parts[4])
        except Exception:
            await send_or_edit(query, 'Ошибка настройки ТОП.')
            return
        try:
            database.set_goal_leaderboard(gid, n)
        except Exception as e:
            await send_or_edit(query, f'Ошибка сохранения: {e}')
            return
        await show_leaderboard_goal_config(query, uid, scope=scope, goal_id=gid)
        return

    if data.startswith('lb_off_'):
        try:
            parts = data.split('_')
            scope = parts[2]
            gid = int(parts[3])
        except Exception:
            await send_or_edit(query, 'Ошибка отключения ТОП.')
            return
        try:
            database.delete_goal_leaderboard(gid)
        except Exception:
            pass
        await show_leaderboard_goal_config(query, uid, scope=scope, goal_id=gid)
        return

    if data.startswith('lb_enter_'):
        try:
            parts = data.split('_')
            scope = parts[2]
            gid = int(parts[3])
        except Exception:
            await send_or_edit(query, 'Ошибка ввода.')
            return
        stx = safe_state(uid)
        stx['mode'] = 'lb_input_n'
        stx['lb_scope'] = scope
        stx['lb_goal_id'] = gid
        await send_or_edit(query, 'Введите число сотрудников для ТОП (например 5).\n\nОтмена — напишите: отмена')
        return

    # RM menu interactions
    if data == 'rm_show_rtps':
        date = datetime.now().strftime('%Y-%m-%d')
        sent_status = database.get_rtp_combined_status_for_all(config.RTP_LIST, date)
        kb = []
        for i, fi in enumerate(config.RTP_LIST):
            status = "✅" if sent_status.get(fi, False) else "❌"
            kb.append([InlineKeyboardButton(f"{fi} {status}", callback_data=f"rm_choose_rtp_{i}")])
        kb.append([InlineKeyboardButton("Объединить все РТП (глобально) и скачать", callback_data='rm_combine_all')])
        kb.append([InlineKeyboardButton("Вернуться в меню", callback_data='return_to_menu')])
        await query.edit_message_text("Список РТП (статус отправки объединённого отчёта):",
                                      reply_markup=InlineKeyboardMarkup(kb))
        return

    if data.startswith('rm_choose_rtp_'):
        # format: rm_choose_rtp_{i}
        try:
            idx = int(data.split('_')[3])
        except Exception:
            await query.edit_message_text("Ошибка выбора.")
            return
        if idx < 0 or idx >= len(config.RTP_LIST):
            await query.edit_message_text("Некорректный индекс.")
            return
        chosen = config.RTP_LIST[idx]
        date = datetime.now().strftime('%Y-%m-%d')
        combined = database.get_rtp_combined(chosen, date)
        if not combined:
            await query.edit_message_text(f"РТП {chosen} не отправлял объединённый отчёт на {date}.",
                                          reply_markup=InlineKeyboardMarkup(
                                              [[InlineKeyboardButton("Назад", callback_data='rm_show_rtps')]]))
            return
        text = f"Объединённый отчёт РТП {chosen} на {date}:\n\n{config.format_report(combined)}"
        kb = [
            [InlineKeyboardButton("📥 Скачать .xlsx", callback_data=f"download_rtp_{idx}")],
            [InlineKeyboardButton("Назад", callback_data='rm_show_rtps')]
        ]
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(kb))
        return

    if data == 'rm_combine_all':
        date = datetime.now().strftime('%Y-%m-%d')
        all_combined = database.get_all_rtp_combined_on_date(date)
        if not all_combined:
            await query.edit_message_text(f"Нет объединённых отчётов от РТП на {date}.",
                                          reply_markup=InlineKeyboardMarkup(
                                              [[InlineKeyboardButton("Назад", callback_data='rm_show_rtps')]]))
            return
        aggregated = {}
        fckp_products = []
        for rtp_fi, rdata in all_combined:
            for k, v in rdata.items():
                if k == 'fckp_products' and isinstance(v, list):
                    fckp_products.extend(v)
                else:
                    try:
                        aggregated[k] = aggregated.get(k, 0) + float(v or 0)
                    except Exception:
                        pass
        aggregated['fckp_products'] = fckp_products
        aggregated['fckp_realized'] = len(fckp_products)
        text = f"Глобальный объединённый отчёт за {date}:\n\n{config.format_report(aggregated)}"
        kb = [
            [InlineKeyboardButton("📥 Скачать глобальный .xlsx", callback_data="download_global")],
            [InlineKeyboardButton("Назад", callback_data='rm_show_rtps')]
        ]
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(kb))
        return

    if data.startswith('download_rtp_'):
        try:
            idx = int(data.split('_')[2])
        except Exception:
            await query.edit_message_text("Ошибка скачивания.")
            return
        if idx < 0 or idx >= len(config.RTP_LIST):
            await query.edit_message_text("Некорректный индекс.")
            return
        rtp_fi = config.RTP_LIST[idx]
        date = datetime.now().strftime('%Y-%m-%d')
        rdata = database.get_rtp_combined(rtp_fi, date)
        if not rdata:
            await query.edit_message_text("Отчёт не найден.")
            return
        rows = []
        for q in config.QUESTIONS:
            rows.append({'key': q['question'], 'value': rdata.get(q['key'], 0)})
        prod_counts = {}
        for p in rdata.get('fckp_products', []):
            prod_counts[p] = prod_counts.get(p, 0) + 1
        for prod in config.FCKP_OPTIONS:
            rows.append({'key': prod, 'value': prod_counts.get(prod, 0)})
        cols = [('key', 'Поле'), ('value', 'Значение')]
        try:
            bio = generate_xlsx_for_report(f"{rtp_fi}_{date}", rows, cols)
            filename = sanitize_filename(f"rtp_{rtp_fi}_{date}.xlsx", default_base="rtp_report")
            await context.bot.send_document(chat_id=uid, document=InputFile(bio, filename=filename))
        except Exception as e:
            await query.edit_message_text(f"Ошибка формирования файла: {e}")
        return

    if data == 'download_global':
        date = datetime.now().strftime('%Y-%m-%d')
        all_combined = database.get_all_rtp_combined_on_date(date)
        rows = []
        for rtp_fi, rdata in all_combined:
            row = {'rtp': rtp_fi}
            for q in config.QUESTIONS:
                row[q['key']] = rdata.get(q['key'], 0)
            row['fckp_count'] = len(rdata.get('fckp_products', []))
            rows.append(row)
        cols = [('rtp', 'RTP')]
        for q in config.QUESTIONS:
            cols.append((q['key'], q['question']))
        cols.append(('fckp_count', 'FCKP count'))
        try:
            bio = generate_xlsx_for_report(f"global_{date}", rows, cols)
            filename = sanitize_filename(f"global_combined_{date}.xlsx", default_base="global_report")
            await context.bot.send_document(chat_id=uid, document=InputFile(bio, filename=filename))
        except Exception as e:
            await query.edit_message_text(f"Ошибка формирования файла: {e}")
        return

    # role_rtp menu (entry)
    if data == 'role_rtp':
        if database.is_user_verified(uid):
            kb = [[InlineKeyboardButton(fi, callback_data=f"choose_rtp_{i}")] for i, fi in enumerate(config.RTP_LIST)]
            kb.append([InlineKeyboardButton("Вернуться в меню", callback_data='return_to_menu')])
            await query.edit_message_text("Выберите ваше ФИ (РТП):", reply_markup=InlineKeyboardMarkup(kb))
            return
        else:
            user_states[uid] = {'mode': 'awaiting_password_for', 'await_role': 'rtp'}
            try:
                await query.edit_message_text("Введите пароль для доступа в раздел РТП:")
            except Exception:
                await query.message.reply_text("Введите пароль для доступа в раздел РТП:")
            return

    if data == 'team_goals_menu':
        owner = database.get_user_name(uid)
        await show_goals_menu(query, uid, scope='team', owner_name=owner, back_cb='rtp_menu')
        return

    # RTP manager actions
    if data == 'rtp_menu':
        await show_manager_menu(query)
        return

    if data == 'rtp_show_reports':
        date = datetime.now().strftime('%Y-%m-%d')
        manager_fi = database.get_user_name(uid)
        employees = database.get_employees(manager_fi)
        reports = database.get_all_reports_on_date(date, manager_fi)
        reported_ids = [u for u, _ in reports]
        text = f"Отчеты на {date}:\n"
        for u_id, name in employees:
            status = '✅' if u_id in reported_ids else '❌'
            text += f"Сотрудник {name or str(u_id)}: {status}\n"
        kb = [[InlineKeyboardButton("Детальный отчет на дату", callback_data='rtp_detailed_reports')],
              [InlineKeyboardButton("Назад", callback_data='rtp_menu')]]
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(kb))
        return

    if data == 'rtp_detailed_reports':
        date = datetime.now().strftime('%Y-%m-%d')
        manager_fi = database.get_user_name(uid)
        reports = database.get_all_reports_on_date(date, manager_fi)
        text = f"Детальные отчеты на {date}:\n\n"
        for u_id, rdata in reports:
            name = database.get_user_name(u_id) or str(u_id)
            text += f"Сотрудник {name}:\n{config.format_report(rdata)}\n\n"
        kb = [[InlineKeyboardButton("Вернуться в меню", callback_data='rtp_menu')]]
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(kb))
        return

    if data == 'rtp_combine_reports':
        date = datetime.now().strftime('%Y-%m-%d')
        manager_fi = database.get_user_name(uid)
        reports = database.get_all_reports_on_date(date, manager_fi)
        if not reports:
            await query.edit_message_text("Нет отчетов на сегодня.", reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton("Назад", callback_data='rtp_menu')]]))
            return
        combined = {}
        fckp_products = []
        for _, r in reports:
            for k, v in r.items():
                if k == 'fckp_products' and isinstance(v, list):
                    fckp_products.extend(v)
                else:
                    try:
                        combined[k] = combined.get(k, 0) + float(v or 0)
                    except Exception:
                        pass
        combined['fckp_products'] = fckp_products
        combined['fckp_realized'] = len(fckp_products)
        text = f"Объединённый отчёт на {date}:\n\n{config.format_report(combined)}\n\n" + config.OPERATIONAL_DEFECTS_BLOCK
        kb = [
            [InlineKeyboardButton("Отправить РМ/МН", callback_data='rtp_send_to_rm')],
            [InlineKeyboardButton("Назад", callback_data='rtp_menu')]
        ]
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(kb))
        return

    if data == 'rtp_send_to_rm':
        manager_fi = database.get_user_name(uid)
        date = datetime.now().strftime('%Y-%m-%d')
        reports = database.get_all_reports_on_date(date, manager_fi)
        if not reports:
            await query.edit_message_text("Нет отчетов для объединения/отправки.", reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton("Назад", callback_data='rtp_menu')]]))
            return
        combined = {}
        fckp_products = []
        for _, r in reports:
            for k, v in r.items():
                if k == 'fckp_products' and isinstance(v, list):
                    fckp_products.extend(v)
                else:
                    try:
                        combined[k] = combined.get(k, 0) + float(v or 0)
                    except Exception:
                        pass
        combined['fckp_products'] = fckp_products
        combined['fckp_realized'] = len(fckp_products)
        database.save_rtp_combined(manager_fi, combined, date)
        await query.edit_message_text("Объединённый отчёт сохранён и доступен РМ/МН.")
        return

    # Editing flow: keep/reselect existing FCKP products
    if data == 'edit_fckp_keep':
        st = safe_state(uid)
        n = int(st.get('pending_fckp_n') or 0)
        existing = st.get('fckp_products') or []
        st['data']['fckp_realized'] = n
        st['data']['fckp_products'] = list(existing)
        st.pop('pending_fckp_n', None)
        st['step'] = st.get('step', 0) + 1
        await ask_next_question(query.message, uid)
        return

    if data == 'edit_fckp_reselect':
        st = safe_state(uid)
        n = int(st.get('pending_fckp_n') or 0)
        st.pop('pending_fckp_n', None)
        if n > 0:
            st['data']['fckp_realized'] = n
            st['fckp_left'] = n
            st['fckp_products'] = []
            kb = [[InlineKeyboardButton(p, callback_data=f"fckp_prod_{p}")] for p in config.FCKP_OPTIONS]
            await send_or_edit(query, f"Выберите оформленный продукт (1/{n}):", reply_markup=InlineKeyboardMarkup(kb))
            return
        await ask_next_question(query.message, uid)
        return

    # FCKP product picking
    if data.startswith('fckp_prod_'):
        prod = data.split('fckp_prod_', 1)[1]
        st = safe_state(uid)
        st.setdefault('fckp_products', [])
        st['fckp_products'].append(prod)
        st['fckp_left'] = st.get('fckp_left', 0) - 1
        left = st.get('fckp_left', 0)
        if left > 0:
            kb = [[InlineKeyboardButton(p, callback_data=f"fckp_prod_{p}")] for p in config.FCKP_OPTIONS]
            try:
                await query.edit_message_text(f"Вы выбрали {prod}. Осталось указать ещё {left} ФЦКП.",
                                              reply_markup=InlineKeyboardMarkup(kb))
            except Exception:
                pass
            return
        else:
            st['data']['fckp_products'] = st.get('fckp_products', [])
            st['data']['fckp_realized'] = len(st.get('fckp_products', []))
            try:
                await query.edit_message_text("Все ФЦКП указаны ✅")
            except Exception:
                pass
            st['step'] = st.get('step', 0) + 1
            await ask_next_question(query.message, uid)
            return

    # download individual user report (RTP view)
    if data.startswith('download_user_'):
        try:
            parts = data.split('_')
            target_uid = int(parts[2])
        except Exception:
            await query.edit_message_text("Ошибка скачивания.")
            return
        date = datetime.now().strftime('%Y-%m-%d')
        rpt = database.get_report(target_uid, date)
        if not rpt:
            await query.edit_message_text("Отчёт не найден.")
            return
        rows = []
        for q in config.QUESTIONS:
            rows.append({'key': q['question'], 'value': rpt.get(q['key'], 0)})
        prod_counts = {}
        for p in rpt.get('fckp_products', []):
            prod_counts[p] = prod_counts.get(p, 0) + 1
        for prod in config.FCKP_OPTIONS:
            rows.append({'key': prod, 'value': prod_counts.get(prod, 0)})
        cols = [('key', 'Поле'), ('value', 'Значение')]
        try:
            bio = generate_xlsx_for_report(f"user_{target_uid}_{date}", rows, cols)
            filename = sanitize_filename(f"user_{target_uid}_{date}.xlsx", default_base="user_report")
            await context.bot.send_document(chat_id=uid, document=InputFile(bio, filename=filename))
        except Exception as e:
            await query.edit_message_text(f"Ошибка формирования файла: {e}")
        return

    if data == 'edit_report':
        await start_edit_report(query, uid)
        return

    if data == 'send_report':
        success, msg_text = await send_personal_report_to_manager(uid, context)
        try:
            if success:
                await query.edit_message_text("Отчёт успешно отправлен руководителю.")
            else:
                await query.edit_message_text(f"Отчет отправлен, но {msg_text}")
        except Exception:
            try:
                await query.message.reply_text("Отчёт отправлен (или произошла ошибка, проверьте лог).")
            except Exception:
                pass
        return

    # fallback
    return


# role selection helper
async def handle_role_selection(query_or_message, user_id, role):
    name = database.get_user_name(user_id)
    if role == 'admin':
        await show_admin_menu(query_or_message)
        return
    if role == 'rtp':
        kb = [[InlineKeyboardButton(fi, callback_data=f"choose_rtp_{i}")] for i, fi in enumerate(config.RTP_LIST)]
        kb.append([InlineKeyboardButton("Вернуться в меню", callback_data='return_to_menu')])
        try:
            await query_or_message.edit_message_text("Выберите ваше ФИ (РТП):", reply_markup=InlineKeyboardMarkup(kb))
        except Exception:
            try:
                await query_or_message.reply_text("Выберите ваше ФИ (РТП):", reply_markup=InlineKeyboardMarkup(kb))
            except Exception:
                pass
        return

    if role == 'rm':
        kb = [[InlineKeyboardButton(fi, callback_data=f"choose_rm_{i}")] for i, fi in enumerate(config.RM_MN_LIST)]
        kb.append([InlineKeyboardButton("Вернуться в меню", callback_data='return_to_menu')])
        try:
            await query_or_message.edit_message_text("Выберите ваше ФИ (РМ/МН):", reply_markup=InlineKeyboardMarkup(kb))
        except Exception:
            try:
                await query_or_message.reply_text("Выберите ваше ФИ (РМ/МН):", reply_markup=InlineKeyboardMarkup(kb))
            except Exception:
                pass
        return

    # MKK flow: ask for name then choose RТП
    if role == 'mkk':
        if name:
            manager_fi = database.get_manager_fi_for_employee(user_id)
            if manager_fi:
                user_states[user_id] = {'mode': role, 'step': 0, 'data': {}, 'editing': False}
                try:
                    await query_or_message.edit_message_text("Роль выбрана. Начинаем заполнение отчёта.")
                except Exception:
                    try:
                        await query_or_message.reply_text("Роль выбрана. Начинаем заполнение отчёта.")
                    except Exception:
                        pass
                await start_filling(query_or_message, user_id)
                return
            else:
                user_states[user_id] = {'mode': role, 'choosing_rtp': True, 'name': name}
                await show_rtp_buttons(query_or_message, "Выберите вашего РТП:")
                return
        else:
            user_states[user_id] = {'mode': role, 'entering_name': True}
            try:
                await query_or_message.edit_message_text("Пожалуйста, введите ваше имя (ФИ):")
            except Exception:
                try:
                    await query_or_message.reply_text("Пожалуйста, введите ваше имя (ФИ):")
                except Exception:
                    pass
            return


async def show_rtp_buttons(query_or_message, text):
    kb = [[InlineKeyboardButton(fi, callback_data=f"choose_rtp_{i}")] for i, fi in enumerate(config.RTP_LIST)]
    kb.append([InlineKeyboardButton("Вернуться в меню", callback_data='return_to_menu')])
    try:
        await query_or_message.reply_text(text, reply_markup=InlineKeyboardMarkup(kb))
    except Exception:
        try:
            await query_or_message.message.reply_text(text, reply_markup=InlineKeyboardMarkup(kb))
        except Exception:
            pass


async def show_manager_menu(q):
    kb = [
        [InlineKeyboardButton("Показать отчеты на дату", callback_data='rtp_show_reports')],
        [InlineKeyboardButton("Детальный отчет на дату", callback_data='rtp_detailed_reports')],
        [InlineKeyboardButton("Объединить и показать отчеты на дату", callback_data='rtp_combine_reports')],
        [InlineKeyboardButton("🎯 Цели команды", callback_data='team_goals_menu')],
        [InlineKeyboardButton("Вернуться в меню", callback_data='return_to_menu')]
    ]
    try:
        await q.edit_message_text("Меню руководителя:", reply_markup=InlineKeyboardMarkup(kb))
    except Exception:
        try:
            await q.message.reply_text("Меню руководителя:", reply_markup=InlineKeyboardMarkup(kb))
        except Exception:
            pass


async def show_rm_home(target, uid: int):
    name = None
    try:
        name = database.get_user_name(uid)
    except Exception:
        name = None
    kb = [
        [InlineKeyboardButton('Список РТП', callback_data='rm_show_rtps')],
        [InlineKeyboardButton('🏢 Управление', callback_data='rm_management')],
        [InlineKeyboardButton('Вернуться в меню', callback_data='return_to_menu')]
    ]
    await send_or_edit(target, f"Меню РМ/МН{f' ({name})' if name else ''}:", reply_markup=InlineKeyboardMarkup(kb))


async def show_rm_management_menu(target, uid: int):
    kb = [
        [InlineKeyboardButton('🎯 Цели ГОСБ', callback_data='gosb_goals_menu')],
        [InlineKeyboardButton('🏆 Лучшие сотрудники', callback_data='gosb_leaderboards_menu')],
        [InlineKeyboardButton('⬅️ Назад', callback_data='rm_menu')]
    ]
    await send_or_edit(target, 'Управление:', reply_markup=InlineKeyboardMarkup(kb))


async def show_leaderboards_menu(target, uid: int, scope: str, owner_name: str = None, back_cb: str = 'return_to_menu'):
    """Configure TOP employees per goal."""
    today = _today_iso()
    try:
        database.cleanup_expired_goals(today)
    except Exception:
        pass

    try:
        goals = database.list_goals('team', owner_name=owner_name,
                                    today=today) if scope == 'team' else database.list_goals('gosb', today=today)
    except Exception:
        goals = []

    title = 'Лучшие сотрудники — Цели ГОСБ' if scope == 'gosb' else f"Лучшие сотрудники — Цели команды ({owner_name})"
    lines = [title + ':']

    kb = []
    if not goals:
        lines.append('Пока нет целей.')
    else:
        for g in goals:
            try:
                n = database.get_goal_leaderboard_top_n(int(g.get('id')))
            except Exception:
                n = 0
            status = f"ТОП: {n}" if n else 'ТОП: выкл'
            lines.append(f"#{g.get('id')} {g.get('title', '')} — {status}")
            kb.append([InlineKeyboardButton(f"⚙️ #{g.get('id')}", callback_data=f"lb_cfg_{scope}_{g.get('id')}")])

    kb.append([InlineKeyboardButton('⬅️ Назад', callback_data=back_cb)])
    await send_or_edit(target, "\n".join(lines).strip(), reply_markup=InlineKeyboardMarkup(kb))


async def show_leaderboard_goal_config(target, uid: int, scope: str, goal_id: int, owner_name: str = None):
    try:
        g = database.get_goal(int(goal_id))
    except Exception:
        g = None
    if not g:
        await send_or_edit(target, 'Цель не найдена.')
        return

    try:
        n = database.get_goal_leaderboard_top_n(int(goal_id))
    except Exception:
        n = 0

    title = g.get('title', '')
    cur = f"Текущее значение ТОП: {n}" if n else 'ТОП сейчас отключён'
    txt = f"""Настройка лучших сотрудников для цели #{goal_id}:

{title}

{cur}

Выберите, сколько сотрудников показывать под этой целью в /start (0 = отключить)."""

    back_to = 'gosb_leaderboards_menu' if scope == 'gosb' else 'team_leaderboards_menu'

    kb = [
        [InlineKeyboardButton('3', callback_data=f'lb_setn_{scope}_{goal_id}_3'),
         InlineKeyboardButton('5', callback_data=f'lb_setn_{scope}_{goal_id}_5'),
         InlineKeyboardButton('10', callback_data=f'lb_setn_{scope}_{goal_id}_10')],
        [InlineKeyboardButton('✏️ Ввести число', callback_data=f'lb_enter_{scope}_{goal_id}')],
        [InlineKeyboardButton('🚫 Отключить ТОП', callback_data=f'lb_off_{scope}_{goal_id}')],
        [InlineKeyboardButton('⬅️ Назад', callback_data=back_to)]
    ]

    await send_or_edit(target, txt, reply_markup=InlineKeyboardMarkup(kb))


async def show_goals_menu(target, uid: int, scope: str, owner_name: str = None, back_cb: str = 'return_to_menu'):
    today = _today_iso()
    try:
        database.cleanup_expired_goals(today)
    except Exception:
        pass

    try:
        goals = database.list_goals('team', owner_name=owner_name,
                                    today=today) if scope == 'team' else database.list_goals('gosb', today=today)
    except Exception:
        goals = []

    title = 'Цели ГОСБ' if scope == 'gosb' else f"Цели команды ({owner_name})"
    lines = [title + ':']

    kb = []
    if not goals:
        lines.append('Пока нет целей.')
    else:
        for g in goals:
            achieved = _compute_goal_achieved(g, today)
            metric = _metric_label(g.get('metric_type'), g.get('metric_key'))
            due = _iso_to_ru(g.get('date_to'))
            a = config.format_value(achieved)
            t = config.format_value(_to_float(g.get('target_value')))
            lines.append((f'''
#{g['id']} {g.get('title', '')}
• Показатель: {metric}
• Прогресс: {a}/{t}
• Срок: до {due}''').strip())
            kb.append([
                InlineKeyboardButton(f"✏️ #{g['id']}", callback_data=f"goal_edit_{scope}_{g['id']}"),
                InlineKeyboardButton('🗑', callback_data=f"goal_del_{scope}_{g['id']}")
            ])

    kb.append([InlineKeyboardButton('➕ Добавить цель', callback_data=f"goal_add_{scope}")])
    if scope == 'team':
        kb.append([InlineKeyboardButton('🏆 Лучшие сотрудники', callback_data='team_leaderboards_menu')])
    kb.append([InlineKeyboardButton('⬅️ Назад', callback_data=back_cb)])

    await send_or_edit(target, "\n\n".join(lines).strip(), reply_markup=InlineKeyboardMarkup(kb))


async def show_goal_edit_menu(target, uid: int, scope: str, goal_id: int, owner_name: str = None):
    try:
        g = database.get_goal(int(goal_id))
    except Exception:
        g = None
    if not g:
        await send_or_edit(target, 'Цель не найдена.')
        return

    today = _today_iso()
    achieved = _compute_goal_achieved(g, today)

    metric = _metric_label(g.get('metric_type'), g.get('metric_key'))
    due = _iso_to_ru(g.get('date_to'))
    frm = _iso_to_ru(g.get('date_from'))
    a = config.format_value(achieved)
    t = config.format_value(_to_float(g.get('target_value')))

    text_msg = f"""Редактирование цели #{g['id']}:

Название: {g.get('title', '')}
Показатель: {metric}
Цель: {t}
Период: {frm} — {due}
Прогресс: {a}/{t}"""

    kb = [
        [InlineKeyboardButton('✏️ Название', callback_data=f"goal_editfield_{scope}_{g['id']}_title")],
        [InlineKeyboardButton('🔗 Показатель', callback_data=f"goal_editfield_{scope}_{g['id']}_metric")],
        [InlineKeyboardButton('🎯 Цель (число)', callback_data=f"goal_editfield_{scope}_{g['id']}_target")],
        [InlineKeyboardButton('📅 Дата начала', callback_data=f"goal_editfield_{scope}_{g['id']}_date_from")],
        [InlineKeyboardButton('📅 Дата окончания', callback_data=f"goal_editfield_{scope}_{g['id']}_date_to")],
        [InlineKeyboardButton('🗑 Удалить', callback_data=f"goal_del_{scope}_{g['id']}")],
        [InlineKeyboardButton('⬅️ Назад', callback_data=f"{scope}_goals_menu")]
    ]
    await send_or_edit(target, text_msg, reply_markup=InlineKeyboardMarkup(kb))


# -----------------------------
# ADMIN UI helpers
# -----------------------------
async def show_admin_menu(target):
    kb = [
        [InlineKeyboardButton("📝 Редактор отчёта МКК", callback_data='admin_edit_questions')],
        [InlineKeyboardButton("👤 Редактор РТП", callback_data='admin_edit_rtps')],
        [InlineKeyboardButton("👥 Редактор сотрудников", callback_data='admin_emp_editor')],
        [InlineKeyboardButton("🔘 Редактор ЦКП", callback_data='admin_edit_fckp')],
        [InlineKeyboardButton("🔑 Пароль РТП", callback_data='admin_set_rtp_password')],
        [InlineKeyboardButton("⬅️ В меню ролей", callback_data='return_to_menu')]
    ]
    await send_or_edit(target, "Панель администратора:", reply_markup=InlineKeyboardMarkup(kb))


async def show_admin_questions_editor(target):
    sync_runtime_config()
    questions = database.get_mkk_questions()

    kb = []
    for i, q in enumerate(questions):
        q_key = q.get('key')
        label = (q.get('question') or '').strip() or f'Вопрос {i + 1}'
        kb.append([InlineKeyboardButton(f"{i + 1}. {label}"[:64], callback_data='noop')])
        # Важно: в callback_data передаём именно ключ вопроса (а не индекс),
        # чтобы редактирование/удаление/перемещение работали корректно.
        kb.append([
            InlineKeyboardButton("⬆️", callback_data=f"admin_q_up_{q_key}"),
            InlineKeyboardButton("⬇️", callback_data=f"admin_q_down_{q_key}"),
            InlineKeyboardButton("✏️", callback_data=f"admin_q_edit_{q_key}"),
            InlineKeyboardButton("🗑", callback_data=f"admin_q_del_{q_key}")
        ])

    kb.append([InlineKeyboardButton("➕ Добавить вопрос", callback_data='admin_q_add')])
    kb.append([InlineKeyboardButton("⬅️ Назад", callback_data='admin_menu')])

    await send_or_edit(target, "Вопросы отчёта МКК:", reply_markup=InlineKeyboardMarkup(kb))


def get_fckp_options():
    """Return current list of CKP/FCKP buttons (from DB setting if exists)."""
    try:
        raw = database.get_setting("fckp_options")
        if raw:
            opts = json.loads(raw)
            if isinstance(opts, list):
                opts = [str(x).strip() for x in opts if str(x).strip()]
                if opts:
                    return opts
    except Exception:
        pass
    return list(getattr(config, "FCKP_OPTIONS", []))


def save_fckp_options(opts: list):
    opts = [str(x).strip() for x in (opts or []) if str(x).strip()]
    database.set_setting("fckp_options", json.dumps(opts, ensure_ascii=False))
    # update runtime
    config.FCKP_OPTIONS = opts
    return opts


async def show_admin_fckp_editor(target):
    opts = get_fckp_options()
    kb = []
    if not opts:
        kb.append([InlineKeyboardButton("Пока нет кнопок ЦКП", callback_data="admin_edit_fckp")])
    for i, opt in enumerate(opts):
        kb.append([InlineKeyboardButton(f"{i + 1}. {opt}", callback_data="admin_edit_fckp")])
        kb.append([
            InlineKeyboardButton("✏️", callback_data=f"admin_fckp_edit_{i}"),
            InlineKeyboardButton("🗑️", callback_data=f"admin_fckp_del_{i}"),
            InlineKeyboardButton("⬆️", callback_data=f"admin_fckp_up_{i}"),
            InlineKeyboardButton("⬇️", callback_data=f"admin_fckp_down_{i}")
        ])
    kb.append([InlineKeyboardButton("➕ Добавить кнопку ЦКП", callback_data="admin_fckp_add")])
    kb.append([InlineKeyboardButton("⬅️ Назад", callback_data="admin_menu")])
    await send_or_edit(target, "Редактор ЦКП (кнопки выбора в ФЦКП):", reply_markup=InlineKeyboardMarkup(kb))


async def show_admin_rtps_editor(target):
    rtps = database.get_rtp_list()
    if not rtps:
        rtps = list(config.RTP_LIST)

    kb = []
    for i, fi in enumerate(rtps):
        label = (fi or '').strip() or f'РТП {i + 1}'
        kb.append([InlineKeyboardButton(f"{i + 1}. {label}"[:64], callback_data='noop')])
        kb.append([
            InlineKeyboardButton("✏️", callback_data=f"admin_rtp_edit_{i}"),
            InlineKeyboardButton("🗑", callback_data=f"admin_rtp_del_{i}")
        ])

    kb.append([InlineKeyboardButton("➕ Добавить РТП", callback_data='admin_rtp_add')])
    kb.append([InlineKeyboardButton("⬅️ Назад", callback_data='admin_menu')])

    await send_or_edit(target, "Список РТП:", reply_markup=InlineKeyboardMarkup(kb))


async def show_admin_employees_rtp_selector(target):
    rtps = database.get_rtp_list()
    if not rtps:
        rtps = list(config.RTP_LIST)

    kb = []
    for i, fi in enumerate(rtps):
        kb.append([InlineKeyboardButton((fi or f"РТП {i + 1}")[:64], callback_data=f"admin_emp_rtp_{i}")])

    kb.append([InlineKeyboardButton("⬅️ Назад", callback_data='admin_menu')])
    await send_or_edit(target, "Выберите РТП, чтобы посмотреть/редактировать сотрудников:",
                       reply_markup=InlineKeyboardMarkup(kb))


async def show_admin_employees_list(target, rtp_idx: int):
    rtps = database.get_rtp_list()
    if not rtps:
        rtps = list(config.RTP_LIST)

    if rtp_idx < 0 or rtp_idx >= len(rtps):
        await send_or_edit(target, "Некорректный индекс РТП.", reply_markup=InlineKeyboardMarkup(
            [[InlineKeyboardButton("⬅️ Назад", callback_data='admin_emp_editor')]]))
        return

    rtp_fi = rtps[rtp_idx]
    employees = database.get_employees(rtp_fi)

    kb = []
    if not employees:
        kb.append([InlineKeyboardButton("(нет сотрудников)", callback_data='noop')])
    else:
        for emp_id, emp_name in employees:
            label = (emp_name or str(emp_id)).strip()
            kb.append([InlineKeyboardButton(label[:64], callback_data='noop')])
            kb.append([
                InlineKeyboardButton("✏️", callback_data=f"admin_emp_edit_{rtp_idx}_{emp_id}"),
                InlineKeyboardButton("🗑", callback_data=f"admin_emp_del_{rtp_idx}_{emp_id}")
            ])

    kb.append([InlineKeyboardButton("⬅️ Назад", callback_data='admin_emp_editor')])
    kb.append([InlineKeyboardButton("🏠 Админ-меню", callback_data='admin_menu')])

    await send_or_edit(target, f"Сотрудники РТП: {rtp_fi}", reply_markup=InlineKeyboardMarkup(kb))


async def show_admin_employee_reassign(target, from_rtp_idx: int, emp_id: int):
    rtps = database.get_rtp_list()
    if not rtps:
        rtps = list(config.RTP_LIST)

    emp_name = database.get_user_name(emp_id) or str(emp_id)

    kb = []
    for i, fi in enumerate(rtps):
        kb.append([InlineKeyboardButton((fi or f"РТП {i + 1}")[:64],
                                        callback_data=f"admin_emp_set_{from_rtp_idx}_{emp_id}_{i}")])

    kb.append([InlineKeyboardButton("🔌 Отвязать от РТП", callback_data=f"admin_emp_unbind_{from_rtp_idx}_{emp_id}")])
    kb.append([InlineKeyboardButton("⬅️ Назад", callback_data=f"admin_emp_rtp_{from_rtp_idx}")])

    await send_or_edit(target, f"Сотрудник: {emp_name}\nВыберите нового РТП:", reply_markup=InlineKeyboardMarkup(kb))


# messages handler
async def message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message
    if not msg:
        return
    uid = msg.from_user.id
    text = (msg.text or "").strip()
    st = user_states.get(uid, {})

    if not st:
        if text.lower() == "вернуться в меню":
            await start(update, context)
            return
        await msg.reply_text("Сессия не запущена. Нажмите /start.")
        return

    if text.lower() == "вернуться в меню":
        user_states.pop(uid, None)
        await start(update, context)
        return

    # Cancel report editing
    if st.get('editing') and text.lower() in ('отмена', 'cancel'):
        st['editing'] = False
        st.pop('pending_fckp_n', None)
        st.pop('fckp_left', None)
        # keep previously saved report
        date = datetime.now().strftime('%Y-%m-%d')
        rpt = database.get_report(uid, date) or st.get('data', {}) or {}
        formatted = config.format_report(rpt)
        await msg.reply_text("Редактирование отменено.")
        await msg.reply_text(f"Текущий отчет:\n{formatted}")
        kb = [[InlineKeyboardButton("Редактировать", callback_data='edit_report')]]
        if st.get('mode') == 'mkk':
            kb[0].append(InlineKeyboardButton("Отправить руководителю", callback_data='send_report'))
        await msg.reply_text("Действия:", reply_markup=InlineKeyboardMarkup(kb))
        return

    # Password entry flow for RM (leader password)
    if st.get('mode') == 'awaiting_password_for':
        await_role = st.get('await_role')
        if text.lower() in ('отмена', 'cancel'):
            user_states.pop(uid, None)
            await msg.reply_text("Отмена. Возврат в меню.", reply_markup=build_main_menu())
            return
        if text == config.ADMIN_PASSWORD:
            database.add_user(uid, await_role)
            database.set_user_verified(uid, 1)
            user_states[uid] = {'mode': await_role, 'step': 0, 'data': {}, 'editing': False}
            await msg.reply_text("Пароль верный. Доступ предоставлен.")
            await handle_role_selection(msg, uid, await_role)
            return
        await msg.reply_text("Неверный пароль. Попробуйте снова")
        return

    # Password entry flow for RTP (separate password)
    if st.get('mode') == 'awaiting_rtp_password':
        if text.lower() in ('отмена', 'cancel'):
            user_states.pop(uid, None)
            await msg.reply_text("Отмена. Возврат в меню.", reply_markup=build_main_menu())
            return
        try:
            ok = (text == database.get_rtp_password())
        except Exception:
            ok = (text == config.ADMIN_PASSWORD)

        if ok:
            # ensure user row exists, mark rtp-verified version
            database.add_user(uid, 'rtp')
            try:
                database.set_user_rtp_verified_version(uid, database.get_rtp_password_version())
            except Exception:
                database.set_user_verified(uid, 1)
            user_states[uid] = {'mode': 'rtp', 'step': 0, 'data': {}, 'editing': False}
            await msg.reply_text("Пароль верный. Доступ предоставлен.")
            await handle_role_selection(msg, uid, 'rtp')
            return

        await msg.reply_text("Неверный пароль РТП. Попробуйте снова")
        return

    # Password entry flow for ADMIN
    if st.get('mode') == 'awaiting_admin_password':
        if text.lower() in ('отмена', 'cancel'):
            user_states.pop(uid, None)
            await msg.reply_text("Отмена. Возврат в меню.", reply_markup=build_main_menu())
            return
        if text == config.ADMIN_PASSWORD:
            database.add_user(uid, 'admin')
            database.set_user_verified(uid, 1)
            user_states[uid] = {'mode': 'admin', 'step': 0, 'data': {}, 'editing': False}
            await msg.reply_text("Пароль верный. Вход в администрирование.")
            await show_admin_menu(msg)
            return
        await msg.reply_text("Неверный пароль. Попробуйте снова")
        return

    # change FI flow
    if st.get('mode') == 'change_fi_enter_name':
        entered_name = text
        st['new_name'] = entered_name
        st['change_flow'] = True
        await show_rtp_buttons(msg, f"Вы ввели имя: {entered_name}\nТеперь выберите вашего РТП из списка:")
        return

    # -----------------------------
    # LEADERBOARD text-input flows
    # -----------------------------
    if st.get('mode') == 'lb_input_n':
        if text.lower() in ('отмена', 'cancel'):
            scope = st.get('lb_scope')
            gid = st.get('lb_goal_id')
            st['mode'] = 'idle'
            await show_leaderboard_goal_config(msg, uid, scope=scope, goal_id=int(gid))
            return
        try:
            n = int(text.strip())
        except Exception:
            await msg.reply_text('Введите целое число, например 5. Или напишите: отмена')
            return
        if n < 0:
            await msg.reply_text('Число не может быть отрицательным. Или напишите: отмена')
            return
        if n > 50:
            await msg.reply_text('Слишком большое число. Максимум 50. Или напишите: отмена')
            return
        scope = st.get('lb_scope')
        gid = int(st.get('lb_goal_id'))
        try:
            database.set_goal_leaderboard(gid, n)
        except Exception as e:
            await msg.reply_text(f'Ошибка сохранения: {e}')
            return
        st['mode'] = 'idle'
        await show_leaderboard_goal_config(msg, uid, scope=scope, goal_id=gid)
        return

    # -----------------------------
    # GOAL text-input flows
    # -----------------------------
    if st.get('mode', '').startswith('goal_'):
        if text.lower() in ('отмена', 'cancel'):
            scope = st.get('goal_scope')
            if scope == 'team':
                owner = database.get_user_name(uid)
                await show_goals_menu(msg, uid, scope='team', owner_name=owner, back_cb='rtp_menu')
            else:
                await show_goals_menu(msg, uid, scope='gosb', back_cb='rm_management')
            return

        mode = st.get('mode')

        if mode == 'goal_add_title':
            st['goal_title'] = text.strip()
            st['mode'] = 'goal_pick_metric'
            st['goal_action'] = 'add'
            await msg.reply_text('Выберите показатель:', reply_markup=_metric_picker_keyboard())
            return

        if mode == 'goal_add_target':
            t = text.replace(',', '.').strip()
            try:
                val = float(t)
            except Exception:
                await msg.reply_text('Введите число, например 30')
                return
            st['goal_target'] = val
            st['mode'] = 'goal_add_date_from'
            await msg.reply_text("Введите дату начала (ДД.ММ.ГГГГ) или 'сегодня':")
            return

        if mode == 'goal_add_date_from':
            try:
                d_from = _parse_date_to_iso(text)
            except Exception:
                await msg.reply_text('Не понял дату. Пример: 20.02.2026')
                return
            st['goal_date_from'] = d_from
            st['mode'] = 'goal_add_date_to'
            await msg.reply_text('Введите дату окончания (ДД.ММ.ГГГГ):')
            return

        if mode == 'goal_add_date_to':
            try:
                d_to = _parse_date_to_iso(text)
            except Exception:
                await msg.reply_text('Не понял дату. Пример: 20.02.2026')
                return
            d_from = st.get('goal_date_from')
            if d_from and d_to < d_from:
                await msg.reply_text('Дата окончания не может быть раньше даты начала. Введите снова:')
                return
            scope = st.get('goal_scope')
            owner = st.get('goal_owner') if scope == 'team' else None
            try:
                database.add_goal(
                    scope=scope,
                    owner_name=owner,
                    title=st.get('goal_title'),
                    metric_type=st.get('goal_metric_type'),
                    metric_key=st.get('goal_metric_key'),
                    target_value=st.get('goal_target', 0),
                    date_from=d_from or _today_iso(),
                    date_to=d_to,
                )
            except Exception as e:
                await msg.reply_text(f'Ошибка создания цели: {e}')
                return

            if scope == 'team':
                owner = database.get_user_name(uid)
                await show_goals_menu(msg, uid, scope='team', owner_name=owner, back_cb='rtp_menu')
            else:
                await show_goals_menu(msg, uid, scope='gosb', back_cb='rm_management')
            return

        if mode == 'goal_edit_title':
            gid = st.get('goal_id')
            try:
                database.update_goal(int(gid), title=text.strip())
            except Exception:
                pass
            await show_goal_edit_menu(msg, uid, scope=st.get('goal_scope'), goal_id=gid,
                                      owner_name=st.get('goal_owner'))
            return

        if mode == 'goal_edit_target':
            gid = st.get('goal_id')
            t = text.replace(',', '.').strip()
            try:
                val = float(t)
            except Exception:
                await msg.reply_text('Введите число, например 30')
                return
            try:
                database.update_goal(int(gid), target_value=val)
            except Exception:
                pass
            await show_goal_edit_menu(msg, uid, scope=st.get('goal_scope'), goal_id=gid,
                                      owner_name=st.get('goal_owner'))
            return

        if mode == 'goal_edit_date_from':
            gid = st.get('goal_id')
            try:
                d_from = _parse_date_to_iso(text)
            except Exception:
                await msg.reply_text('Не понял дату. Пример: 20.02.2026')
                return
            g = database.get_goal(int(gid)) or {}
            d_to = g.get('date_to')
            if d_to and d_to < d_from:
                await msg.reply_text('Дата начала не может быть позже даты окончания. Введите снова:')
                return
            try:
                database.update_goal(int(gid), date_from=d_from)
            except Exception:
                pass
            await show_goal_edit_menu(msg, uid, scope=st.get('goal_scope'), goal_id=gid,
                                      owner_name=st.get('goal_owner'))
            return

        if mode == 'goal_edit_date_to':
            gid = st.get('goal_id')
            try:
                d_to = _parse_date_to_iso(text)
            except Exception:
                await msg.reply_text('Не понял дату. Пример: 20.02.2026')
                return
            g = database.get_goal(int(gid)) or {}
            d_from = g.get('date_from')
            if d_from and d_to < d_from:
                await msg.reply_text('Дата окончания не может быть раньше даты начала. Введите снова:')
                return
            try:
                database.update_goal(int(gid), date_to=d_to)
            except Exception:
                pass
            await show_goal_edit_menu(msg, uid, scope=st.get('goal_scope'), goal_id=gid,
                                      owner_name=st.get('goal_owner'))
            return

    # -----------------------------
    # ADMIN text-input flows
    # -----------------------------
    if st.get('mode') == 'admin_questions_add':
        try:
            database.add_mkk_question(text)
        except Exception:
            pass
        sync_runtime_config()
        user_states[uid] = {'mode': 'admin_edit_questions'}
        await show_admin_questions_editor(msg)
        return

    if st.get('mode') == 'admin_q_edit':
        q_key = st.get('q_key')
        if q_key:
            try:
                database.update_mkk_question(q_key, text)
            except Exception:
                pass
        sync_runtime_config()
        user_states[uid] = {'mode': 'admin_edit_questions'}
        await show_admin_questions_editor(msg)
        return

    if st.get('mode') == 'admin_rtp_add':
        try:
            ok = database.add_rtp(text)
        except Exception:
            ok = False
        sync_runtime_config()
        user_states[uid] = {'mode': 'admin_edit_rtps'}
        if not ok:
            await msg.reply_text("Не удалось добавить РТП (возможно, такое ФИ уже есть).")
        await show_admin_rtps_editor(msg)
        return

    if st.get('mode') == 'admin_rtp_edit':
        old_name = st.get('old_name')
        if old_name:
            try:
                ok = database.update_rtp(old_name, text)
            except Exception:
                ok = False
            if not ok:
                await msg.reply_text("Не удалось переименовать РТП (возможно, такое ФИ уже есть).")
        sync_runtime_config()
        user_states[uid] = {'mode': 'admin_edit_rtps'}
        await show_admin_rtps_editor(msg)
        return

    if st.get('mode') == 'admin_fckp_add':
        opts = get_fckp_options()
        opts.append(text.strip())
        save_fckp_options(opts)
        user_states[uid] = {'mode': 'admin_edit_fckp'}
        await show_admin_fckp_editor(msg)
        return

    if st.get('mode') == 'admin_fckp_edit':
        idx = st.get('fckp_idx')
        opts = get_fckp_options()
        if isinstance(idx, int) and 0 <= idx < len(opts):
            opts[idx] = text.strip()
            save_fckp_options(opts)
        user_states[uid] = {'mode': 'admin_edit_fckp'}
        await show_admin_fckp_editor(msg)
        return

    if st.get('mode') == 'admin_set_rtp_password':
        try:
            ok = database.set_rtp_password(text)
        except Exception:
            ok = False
        if ok:
            await msg.reply_text("Пароль РТП обновлён. Всем РТП потребуется ввести новый пароль при следующем входе.")
        else:
            await msg.reply_text("Не удалось обновить пароль (пустое значение?).")
        user_states[uid] = {'mode': 'admin', 'step': 0, 'data': {}, 'editing': False}
        await show_admin_menu(msg)
        return

    # Registration flows (MKK name entering)
    if st.get('entering_name'):
        name = text
        role = st.get('mode', 'idle')
        st['name'] = name
        st.pop('entering_name', None)
        database.add_user(uid, 'mkk' if role == 'mkk' else role, name)
        if role == 'mkk':
            st['choosing_rtp'] = True
            await show_rtp_buttons(update, "Выберите вашего РТП:")
        else:
            await msg.reply_text("Выберите ваше ФИ из списка кнопок.")
        return

    if st.get('choosing_rtp'):
        await msg.reply_text("Пожалуйста, выберите РТП из списка кнопок.")
        return

    # Now questionnaire: accept floats (and ints)
    if 'step' not in st:
        return

    step = st['step']
    if step < len(config.QUESTIONS):
        q = config.QUESTIONS[step]
        # Accept float-like input (allow comma)
        t = text.replace(',', '.')
        try:
            if t == '':
                val = 0.0
            else:
                val = float(t)
        except Exception:
            await msg.reply_text("Пожалуйста, введите число (можно дробное, например 1.5 либо 0,7).")
            return

        # special handling for fckp_realized (asks for product choices)
        if q['key'] == 'fckp_realized':
            n = int(val)

            # If editing and count is unchanged, offer keep/reselect products
            existing = st.get('fckp_products') if isinstance(st.get('fckp_products'), list) else st.get('data', {}).get(
                'fckp_products', [])
            if st.get('editing') and isinstance(existing, list) and existing and len(existing) == n and n > 0:
                st['pending_fckp_n'] = n
                kb = [[
                    InlineKeyboardButton("Оставить текущие", callback_data="edit_fckp_keep"),
                    InlineKeyboardButton("Выбрать заново", callback_data="edit_fckp_reselect")
                ]]
                await msg.reply_text("Количество ФЦКП не изменилось. Оставить текущий список продуктов?",
                                     reply_markup=InlineKeyboardMarkup(kb))
                return

            st['data'][q['key']] = n
            if n > 0:
                st['fckp_left'] = n
                st['fckp_products'] = []
                kb = [[InlineKeyboardButton(p, callback_data=f"fckp_prod_{p}")] for p in config.FCKP_OPTIONS]
                await msg.reply_text(f"Вы указали {n} ФЦКП. Выберите оформленный продукт (1/{n}):",
                                     reply_markup=InlineKeyboardMarkup(kb))
                return
            else:
                # clear products
                st['data']['fckp_products'] = []
                st.pop('fckp_products', None)
                st.pop('fckp_left', None)
                st['step'] += 1
                await ask_next_question(msg, uid)
                return
        else:
            st['data'][q['key']] = str(val)
            st['step'] += 1
            await ask_next_question(msg, uid)
            return
    else:
        await msg.reply_text("Опрос завершён. Для возврата в меню нажмите 'Вернуться в меню' или /start.")
        return


async def ask_next_question(msgobj, uid):
    st = safe_state(uid)
    step = st.get('step', 0)
    if step < len(config.QUESTIONS):
        q = config.QUESTIONS[step]
        current = st.get('data', {}).get(q['key'], '')
        try:
            await msgobj.reply_text(f"{q['question']} {f'(текущее: {current})' if current != '' else ''}")
        except Exception:
            try:
                await msgobj.message.reply_text(f"{q['question']} {f'(текущее: {current})' if current != '' else ''}")
            except Exception:
                pass
    else:
        await finish_report(msgobj, uid)


async def start_filling(query_or_message, uid, editing=False):
    st = safe_state(uid)
    st['editing'] = editing
    st['step'] = 0
    if not editing:
        st['data'] = {}
    try:
        await query_or_message.edit_message_text("Начинаем заполнение отчёта.")
    except Exception:
        try:
            await query_or_message.reply_text("Начинаем заполнение отчёта.")
        except Exception:
            pass
    await ask_next_question(query_or_message, uid)


async def start_edit_report(query_or_message, uid):
    """Start step-by-step editing of the saved report for today."""
    date = datetime.now().strftime('%Y-%m-%d')
    rpt = database.get_report(uid, date) or {}
    st = safe_state(uid)
    # keep current role/mode, but switch to editing
    st['editing'] = True
    st['step'] = 0
    st['data'] = dict(rpt) if isinstance(rpt, dict) else {}
    # keep existing product list if any
    if isinstance(st['data'].get('fckp_products'), list):
        st['fckp_products'] = list(st['data'].get('fckp_products'))
    else:
        st.pop('fckp_products', None)
    st.pop('fckp_left', None)

    await send_or_edit(query_or_message,
                       "Редактирование отчёта. Введите значения по пунктам (можно оставить прежнее, введя то же число).")
    await ask_next_question(query_or_message, uid)


async def finish_report(msgobj, uid):
    st = safe_state(uid)
    data = st.get('data', {}) or {}
    if 'fckp_products' in st and st.get('fckp_products'):
        data['fckp_products'] = st.get('fckp_products')
        data['fckp_realized'] = len(st.get('fckp_products'))
    # ensure all questions present
    for q in config.QUESTIONS:
        data.setdefault(q['key'], 0)
    try:
        if st.get('mode') != 'idle':
            database.save_report(uid, data)
    except Exception as e:
        print("DB save_report error:", e)
    formatted = config.format_report(data)
    try:
        await msgobj.reply_text(f"Итоговый отчет:\n{formatted}")
    except Exception:
        try:
            await msgobj.message.reply_text(f"Итоговый отчет:\n{formatted}")
        except Exception:
            pass
    # Final actions menu (no change_info here per request)
    kb = [
        [InlineKeyboardButton("Редактировать", callback_data='edit_report')]
    ]
    if st.get('mode') == 'mkk':
        kb[0].insert(1, InlineKeyboardButton("Отправить руководителю", callback_data='send_report'))
    try:
        await msgobj.reply_text("Действия:", reply_markup=InlineKeyboardMarkup(kb))
    except Exception:
        pass


async def send_personal_report_to_manager(uid, context):
    date = datetime.now().strftime('%Y-%m-%d')
    rpt = database.get_report(uid, date)
    if not rpt:
        return False, "Отчёт не найден"
    formatted = config.format_report(rpt)
    name = database.get_user_name(uid) or str(uid)
    manager_fi = database.get_manager_fi_for_employee(uid)
    if not manager_fi:
        return False, "Руководитель не привязан"
    manager_id = database.get_manager_id_by_fi(manager_fi)
    if not manager_id:
        return False, f"руководитель {manager_fi} не найден в системе"
    try:
        await context.bot.send_message(chat_id=manager_id, text=f"Отчёт от сотрудника {name} на {date}:\n{formatted}")
        return True, "Отчёт отправлен"
    except Exception as e:
        print("send_personal_report_to_manager error:", e)
        return False, "Ошибка отправки"


async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    print("Error:", context.error)


async def set_commands(app):
    try:
        await app.bot.set_my_commands([BotCommand("start", "Начать работу с ботом")])
    except Exception as e:
        print("set_commands error:", e)


if __name__ == '__main__':
    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler('start', start))
    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, message_handler))
    app.add_error_handler(error_handler)
    asyncio.get_event_loop().run_until_complete(set_commands(app))
    print("Bot started")
    app.run_polling()
